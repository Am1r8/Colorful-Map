print("This code is developed by AlPHA")

import osmnx as ox
from twython import Twython
from instabot import Bot
import openpyxl
from time import sleep
from PIL import Image
import os
from auth import (
    consumer_key,
    consumer_secret,
    access_token,
    access_token_secret
)




# infos
lat = 0
lng = 0
picname = ""
instabot = Bot()

# Your instagram Password and Username 
instabot.login(username = "", password = "")
wrkbk  = openpyxl.load_workbook('worldcities_clean.xlsx')
sh = wrkbk.active


def making_map():
    point = (lat, lng)
    G = ox.graph_from_point(point, dist=10000, retain_all=True, simplify = True, network_type='all')


    u = []
    v = []
    key = []
    data = []
    for uu, vv, kkey, ddata in G.edges(keys=True, data=True):
        u.append(uu)
        v.append(vv)
        key.append(kkey)
        data.append(ddata)    

    # List to store colors
    roadColors = []
    roadWidths = []

    for item in data:
        if "length" in item.keys():
            if item["length"] <= 100:
                linewidth = 0.10
                color = "#a6a6a6" 
                
            elif item["length"] > 100 and item["length"] <= 200:
                linewidth = 0.15
                color = "#676767"
                
            elif item["length"] > 200 and item["length"] <= 400:
                linewidth = 0.25
                color = "#454545"
                
            elif item["length"] > 400 and item["length"] <= 800:
                color = "#d5d5d5"
                linewidth = 0.35
            else:
                color = "#ededed"
                linewidth = 0.45
        else:
            color = "#a6a6a6"
            linewidth = 0.10
                
        roadColors.append(color)
        roadWidths.append(linewidth)
                


    latitude = 40.4381311
    longitude = -3.8196194



    bgcolor = "#061529"

    fig, ax = ox.plot_graph(G, node_size=0,figsize=(5, 5), 
                            dpi = 300,bgcolor = bgcolor,
                            save = False, edge_color=roadColors,
                            edge_linewidth=roadWidths, edge_alpha=1, show=False)

    fig.tight_layout(pad=0)
    fig.savefig(picname+".png", dpi=300, bbox_inches='tight', format="png", facecolor=fig.get_facecolor(), transparent=False)



# twitter
def twitter():
    twitter = Twython(consumer_key, consumer_secret, access_token, access_token_secret)

    message = "This is Map of " + picname + " This picture is made by a computer and it's written With python. #city #python #programming #computers #ai #AlPHA"
    image = open(picname+".png", 'rb')
    response = twitter.upload_media(media=image)
    media_id = [response['media_id']]
    twitter.update_status(status=message, media_ids=media_id)


#instagram
def inst():
    im = Image.open(picname+".png")
    rgb_im = im.convert('RGB')
    rgb_im.save(picname+'.jpg')
    instabot.upload_photo(picname+".jpg", caption ="This is Map of " + picname + " This picture is made by a computer and it's written With python. #city #python #programming #computers #ai #AlPHA")




for i in range(1, sh.max_row+1):
    cities = []
      
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
        cities.append(cell_obj.value)
        sleep(1)
    picname = cities[0] + " " + cities[3]
    lat = cities[1]
    lng = cities[2]
    print(lat)
    print(lng)
    making_map()
    twitter()
    inst()
    sleep(18000)
    os.remove(picname+".png")
    sleep(5)